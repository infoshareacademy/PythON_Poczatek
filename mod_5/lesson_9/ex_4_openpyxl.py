from openpyxl import load_workbook


def run_example():

    workbook = load_workbook(filename="my-doc.xlsx")
    print(workbook.sheetnames)
    first_sheet = workbook.active
    print(first_sheet.title)

    print(first_sheet["A1"].value, first_sheet["B1"].value)
    print(first_sheet["A2"].value, first_sheet["B2"].value)

    first_sheet["A1"].value = "Hello!"
    print(first_sheet["A1"].value)

    workbook.save(filename="updated-doc.xlsx")


if __name__ == "__main__":
    run_example()
